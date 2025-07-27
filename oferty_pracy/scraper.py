import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import urllib.parse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def get_offer_details(offer_url):
    """Fetches and parses the details from a single job offer page."""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        response = requests.get(offer_url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        description_div = soup.find("div", class_="description__text")
        if description_div:
            description = description_div.get_text(separator='\n', strip=True)
            return description
        return "Description not found."
    except requests.exceptions.RequestException as e:
        print(f"Error fetching offer details from {offer_url}: {e}")
        return "Could not fetch details."
    except Exception as e:
        print(f"An error occurred while parsing details from {offer_url}: {e}")
        return "Could not parse details."

def get_offers(keyword, location, start=0):
    keyword_encoded = urllib.parse.quote(keyword)
    location_encoded = urllib.parse.quote(location)
    url = f"https://www.linkedin.com/jobs/search/?keywords={keyword_encoded}&location={location_encoded}&start={start}"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching the search results page: {e}")
        return []

    soup = BeautifulSoup(response.text, "html.parser")
    job_listings = soup.find_all("div", class_="base-card")

    offers = []
    for job in job_listings:
        try:
            title = job.find("h3", class_="base-search-card__title").get_text(strip=True)
            company = job.find("h4", class_="base-search-card__subtitle").get_text(strip=True)
            location = job.find("span", class_="job-search-card__location").get_text(strip=True)
            link = job.find("a", class_="base-card__full-link")["href"]

            if link.startswith('/'):
                link = "https://www.linkedin.com" + link

            offers.append({
                "title": title,
                "company": company,
                "location": location,
                "link": link,
                "description": ""
            })
        except AttributeError:
            continue

    return offers

if __name__ == "__main__":
    keyword_input = input("Enter the job title/keyword to search for (e.g., Python Developer): ")
    location_input = input("Enter the location (e.g., Warsaw): ")
    try:
        num_pages_to_scrape = int(input("How many pages of results would you like to scrape? "))
    except ValueError:
        print("Invalid number of pages. Defaulting to 1.")
        num_pages_to_scrape = 1

    all_offers = []
    for page_num in range(num_pages_to_scrape):
        print(f"Scraping page {page_num + 1}...")
        try:
            offers_on_page = get_offers(keyword_input, location_input, start=page_num * 25)
            if not offers_on_page:
                print("No more offers found.")
                break
            
            # Filter offers to only include those with the exact keyword in the title
            filtered_offers = [
                offer for offer in offers_on_page
                if keyword_input.lower() in offer['title'].lower()
            ]

            for offer in filtered_offers:
                print(f"  Fetching details for: {offer['title']} at {offer['company']}")
                offer['description'] = get_offer_details(offer['link'])
                time.sleep(1)

            all_offers.extend(filtered_offers)
            print(f"Page {page_num + 1} scraped. Found {len(filtered_offers)} matching offers. Total collected: {len(all_offers)}")
            time.sleep(2)

        except Exception as e:
            print(f"An error occurred on page {page_num + 1}: {e}")
            continue

    if all_offers:
        df = pd.DataFrame(all_offers)
        output_filename = 'linkedin_jobs.xlsx'
        
        # Save the DataFrame to an Excel file
        df.to_excel(output_filename, index=False, engine='openpyxl')

        # --- Excel Formatting ---
        try:
            wb = load_workbook(output_filename)
            ws = wb.active

            # 1. Make headers bold
            bold_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold_font

            # 2. Adjust column widths and wrap text for description
            for i, column_cells in enumerate(ws.columns):
                column_letter = get_column_letter(column_cells[0].column)
                
                # For the description column, enable text wrapping and set a fixed width
                if ws.cell(row=1, column=i+1).value == 'description':
                    ws.column_dimensions[column_letter].width = 80
                    for cell in column_cells:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    # For other columns, adjust width based on content length
                    # Use a reasonable max width to avoid excessively wide columns
                    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                    ws.column_dimensions[column_letter].width = min(length + 2, 60)

            # Save the workbook with formatting
            wb.save(output_filename)
            print(f"Successfully saved and formatted {len(all_offers)} job offers to {output_filename}")
        except Exception as e:
            print(f"Could not apply formatting to the Excel file. Error: {e}")
            print(f"Data is still saved in {output_filename} without extra formatting.")

    else:
        print("No job offers were found to save.")